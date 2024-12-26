from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


def create_wb() -> Workbook:
    """
    Creates a new Workbook object.

    Returns:
        Workbook: A new Workbook instance.
    """
    return Workbook()


def load_wb(filename: str, data_only: bool = False) -> Workbook:
    """
    Loads an existing workbook from a file.

    Args:
        filename (str): The path to the Excel file.
        data_only (bool): If True, only read the cell values, not the formulas. Defaults to False.

    Returns:
        Workbook: The loaded Workbook instance.
    """
    return load_workbook(filename=filename, data_only=data_only)


def apply_font(
    cell: Cell,
    bold: bool = False,
    color: str = "00000000",
    name: str = "Arial",
    size: int = 10,
) -> None:
    """
    Applies font styles to a cell.

    Args:
        cell (Cell): The cell to apply font styles to.
        bold (bool): Whether to apply bold formatting. Defaults to False.
        color (str): The color of the font in hexadecimal format. Defaults to black.
        name (str): The name of the font. Defaults to "Arial".
        size (int): The font size. Defaults to 10.
    """
    cell.font = Font(bold=bold, color=color, name=name, size=size)


def apply_fill(cell: Cell, color: str) -> None:
    """
    Applies a solid background color to a cell.

    Args:
        cell (Cell): The cell to apply the background color to.
        color (str): The background color in hexadecimal format.
    """
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def get_side(style: str = "medium", color: str = "000000") -> Side:
    """
    Creates a Side object for borders.

    Args:
        style (str): The style of the border (e.g., "thin", "medium", "thick"). Defaults to "medium".
        color (str): The color of the border in hexadecimal format. Defaults to black.

    Returns:
        Side: The created Side object.
    """
    return Side(style=style, color=color)


def get_border(
    bottom: Side | None = None,
    left: Side | None = None,
    right: Side | None = None,
    top: Side | None = None,
) -> Border:
    """
    Creates a Border object for cells.

    Args:
        bottom (Side): The bottom border.
        left (Side): The left border.
        right (Side): The right border.
        top (Side): The top border.

    Returns:
        Border: The created Border object.
    """
    return Border(bottom=bottom, left=left, right=right, top=top)


def apply_center_alignment(cell: Cell) -> Alignment:
    """
    Applies center alignment to a cell.

    Args:
        cell (Cell): The cell to apply center alignment to.

    Returns:
        Alignment: The applied Alignment object.
    """
    cell.alignment = Alignment(horizontal="center")
    return cell.alignment


def apply_header_cell(cell: Cell) -> None:
    """
    Applies header cell formatting to a cell, including center alignment, bottom border, and bold font.

    Args:
        cell (Cell): The cell to apply header cell formatting to.
    """
    cell.alignment = apply_center_alignment(cell)
    cell.border = get_border(bottom=get_side())
    cell.font = Font(bold=True)


def add_comment(cell: Cell, comment: str, author: str) -> None:
    """
    Adds a comment to a cell.

    Args:
        cell (Cell): The cell to add the comment to.
        comment (str): The text of the comment.
        author (str): The author of the comment.
    """
    cell.comment = Comment(comment, author)


def apply_number_format(cell: Cell, format: str = "R$ #,##0.00") -> None:
    """
    Applies a number format to a cell.

    Args:
        cell (Cell): The cell to apply the number format to.
        format (str): The number format string.
        Defaults to "R$ #,##0.00" BRL Corrency.

    **Examples:** "0.00%", "0%", "#,##0", "MM/YYYY", "MM/DD/YYYY",
    """
    cell.number_format = format
